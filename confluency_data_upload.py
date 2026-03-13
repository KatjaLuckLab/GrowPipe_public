#!/usr/bin/env python

# this script uploads Incucyte data and plate layout info to MySQL into their respective tables: Confluency Dataset, Condition ID, and Meta data table
# this script also updates the database with new Incucyte data, by filtering out conditions that already exist

import pandas as pd
import numpy as np
import os
import db_utils
import glob
from itertools import chain
import openpyxl
from openpyxl import load_workbook
import mysql.connector

# import necessary functions
import confluency_upload_functions

#import Incucyte data into mysql

#######################################################################################################################################################################

# Incucyte Confluency Dataset

# upload raw Confluency data and update Confluency Dataset in MySQL using overview of Incucyte experiments excel file to parse for recently uploaded data

# read file

Path = '/'

OverviewOfIncucyteExperimentsFile = Path + 'Overview_Incucyte_Experiments.xlsx'

OverviewOfIncucyteExperimentsDF =pd.read_excel(OverviewOfIncucyteExperimentsFile)

ExcelWorkbook = openpyxl.load_workbook(OverviewOfIncucyteExperimentsFile)

# get experiment, project, plate IDs and dates for each plate that is in excel spreadsheet but has not yet been uploaded to MySQL
# each row that is parsed, its "In MySQL?" column will be updated 0 -> 1

ExperimentNames = []
ProjectNames = []
PlateNames = []
Dates = []

i = 1
# and row[7]==0
for row in OverviewOfIncucyteExperimentsDF.itertuples():
	if row[6]==1 and row[8]==1 and row[9]==0 and row[10]==1:
          i +=1
          ExperimentNames.append(row[1])
          ProjectNames.append(row[2])
          PlateNames.append(row[3])
          Dates.append(row[5])
          ExcelWorkbook['Tabelle1'].cell(column=7, row = i).value=1


ExcelWorkbook.save(OverviewOfIncucyteExperimentsFile)

# want to save data files and platelayout files that are to be uploaded
ListOfIncucyteDataFilenames = []
ListofIncucytePlatelayouts = []


ListOfIncucyteDataFilenames = incu_mysql_upload_functions.getIncucyteDataFilenames(PlateNames, ExperimentNames, ProjectNames, Dates, ListOfIncucyteDataFilenames)

ListofIncucytePlatelayouts = incu_mysql_upload_functions.getIncucytePlatelayouts(PlateNames, ExperimentNames, ProjectNames, Dates, ListofIncucytePlatelayouts)


connect = mysql.connector.connect(
  host="",
  user="",
  password="",
  database=""
)

cursor = connect.cursor()

IncucyteData = []

IncucyteData = incu_mysql_upload_functions.loadIncucyteData(ListOfIncucyteDataFilenames, IncucyteData)


#insert Incucyte data into Incucyte data table in MySQL
sqlInsert1 = """insert into BAF_project.Incucyte_Confluency_Dataset (project_ID,plate_ID,timepoint,well_ID,confluency) values (%s,%s,%s,%s,%s)"""



# check if Confluency table already exists in MySQL
ConfluencyTableExistsInMySQL =incu_mysql_upload_functions.checkTableExists('Incucyte_Confluency_Dataset', cursor)

if ConfluencyTableExistsInMySQL == True:
    cursor.executemany(sqlInsert1,IncucyteData)

if ConfluencyTableExistsInMySQL == False:
    create_statement1=  """CREATE TABLE BAF_project.Incucyte_Confluency_Dataset (
                            project_ID varchar(50) NOT NULL,
                            plate_ID varchar(25) NOT NULL,
                            timepoint double NOT NULL,
                            well_ID varchar(25) NOT NULL,
                            confluency double NOT NULL,
                            exclude int(1) DEFAULT 0,
                            PRIMARY KEY (project_ID,plate_ID,well_ID,timepoint))"""
    cursor.execute(create_statement1)

    cursor.executemany(sqlInsert1,IncucyteData)

#commit to the db
connect.commit()

#######################################################################################################################################################################

# Incucyte Condition ID Table

# first check whether Condition ID table already exists in MySQL


condIDTableExistsInMySQL =incu_mysql_upload_functions.checkTableExists('Incucyte_Condition_Table', cursor)


if condIDTableExistsInMySQL == True:

    # get cell line and treatment info and condition IDs from existing Incucyte Meta Data table from MySQL
    query1 = 'select * from BAF_project.Incucyte_Condition_Table'
    cursor.execute(query1)
    IncucyteConditionTableTuple = cursor.fetchall()

    # converting IncucyteConditionTableTuple to ExistingConditionTableContents
    ExistingConditionTableContents = pd.DataFrame(list(IncucyteConditionTableTuple), columns=['condition_ID', 'cell_line', 
                        'cell_line_modifications', 'cell_number', 'treatment_0', 'concentration_0', 'treatment_timepoint_0', 
                        'treatment_duration_0','treatment_1', 'concentration_1', 'treatment_timepoint_1', 
                        'treatment_duration_1','treatment_2', 'concentration_2', 'treatment_timepoint_2', 
                        'treatment_duration_2','treatment_3', 'concentration_3', 'treatment_timepoint_3', 
                        'treatment_duration_3'])  

    # create condition_ID
    # create empty dataframe 
    IncucyteMetaData = pd.DataFrame()

    IncucyteMetaData = incu_mysql_upload_functions.loadIncucyteMetaData(ListofIncucytePlatelayouts, IncucyteMetaData)

    # remove condition duplicates for all columns
    IncucyteMetaData.drop_duplicates(ignore_index=True,inplace = True)

    IncucyteMetaData= IncucyteMetaData.replace(np.nan,None)

    # remove rows where well was empty
    IncucyteMetaData.dropna(subset=['cell_line_modifications'], inplace=True)

    # need to combine existing conditions table w/ input_data and assign the same conditions to the same condition_ID
    ExistingAndNewConditionsOverlapDF=pd.merge(ExistingConditionTableContents,IncucyteMetaData,how = 'outer', on=['cell_line', 
                        'cell_line_modifications', 'cell_number', 'treatment_0', 'concentration_0', 'treatment_timepoint_0', 
                        'treatment_duration_0','treatment_1', 'concentration_1', 'treatment_timepoint_1', 
                        'treatment_duration_1','treatment_2', 'concentration_2', 'treatment_timepoint_2', 
                        'treatment_duration_2','treatment_3', 'concentration_3', 'treatment_timepoint_3', 
                        'treatment_duration_3'])

    ExistingAndNewConditionsOverlapDF= ExistingAndNewConditionsOverlapDF.replace(np.nan,None)

    IndexValue = len(IncucyteConditionTableTuple)

    k = 0

    # create empty series
    SeriesOfConditions = pd.Series([],dtype=pd.StringDtype())

    ExistingAndNewConditionsOverlapDF = incu_mysql_upload_functions.createAdditionalConditionIDs(ExistingAndNewConditionsOverlapDF, SeriesOfConditions, IndexValue, k)

    ExistingAndNewConditionsDF = pd.concat([ExistingAndNewConditionsOverlapDF,ExistingConditionTableContents]).drop_duplicates(keep=False)

    # reset index numbering
    ExistingAndNewConditionsDF=ExistingAndNewConditionsDF.reset_index()

    # remove unnecessary column
    ExistingAndNewConditionsDF = ExistingAndNewConditionsDF.drop(['index'], axis =1)

    # insert query
    sqlInsert2 = """INSERT INTO BAF_project.Incucyte_Condition_Table (condition_ID, cell_line, 
                        cell_line_modifications, cell_number, treatment_0, concentration_0, treatment_timepoint_0, 
                        treatment_duration_0,treatment_1, concentration_1, treatment_timepoint_1, 
                        treatment_duration_1,treatment_2, concentration_2, treatment_timepoint_2, 
                        treatment_duration_2,treatment_3, concentration_3, treatment_timepoint_3, 
                        treatment_duration_3) 
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

    # get length of dataframe
    TableSize = ExistingAndNewConditionsDF.shape[0]

    # update condition table in database; insert values from ExistingAndNewConditionsDF row-wise into the MySQL table
    for i in range(TableSize):
        cursor.execute(sqlInsert2,tuple(ExistingAndNewConditionsDF.loc[i]))


# if there is no existing Condition ID table in MySQL, one will be created
if condIDTableExistsInMySQL == False:
    create_statement=  """CREATE TABLE BAF_project.Incucyte_Condition_Table (
                            condition_ID varchar(45) NOT NULL,
                            cell_line varchar(45) NOT NULL,
                            cell_line_modifications varchar(45) NOT NULL,
                            cell_number int(8) NOT NULL,
                            treatment_0 varchar(45) DEFAULT NULL,
                            concentration_0 varchar(45) DEFAULT NULL,
                            treatment_timepoint_0 varchar(45) DEFAULT NULL,
                            treatment_duration_0 varchar(45) DEFAULT NULL,
                            treatment_1 varchar(45) DEFAULT NULL,
                            concentration_1 varchar(45) DEFAULT NULL,
                            treatment_timepoint_1 varchar(45) DEFAULT NULL,
                            treatment_duration_1 varchar(45) DEFAULT NULL,
                            treatment_2 varchar(45) DEFAULT NULL,
                            concentration_2 varchar(45) DEFAULT NULL,
                            treatment_timepoint_2 varchar(45) DEFAULT NULL,
                            treatment_duration_2 varchar(45) DEFAULT NULL,
                            treatment_3 varchar(45) DEFAULT NULL,
                            concentration_3 varchar(45) DEFAULT NULL,
                            treatment_timepoint_3 varchar(45) DEFAULT NULL,
                            treatment_duration_3 varchar(45) DEFAULT NULL,
                            PRIMARY KEY (condition_ID))"""
    cursor.execute(create_statement)

    # create condition_ID

    # create empty dataframe 
    IncucyteMetaData = pd.DataFrame()


    IncucyteMetaData = incu_mysql_upload_functions.loadIncucyteMetaData(ListofIncucytePlatelayouts, IncucyteMetaData)

    # remove condition duplicates for all columns
    IncucyteMetaData.drop_duplicates(ignore_index=True,inplace = True)

    IncucyteMetaData= IncucyteMetaData.replace(np.nan,None)

    # remove condition duplicates for all columns
    IncucyteMetaData.drop_duplicates(ignore_index=True,inplace = True)

    # remove rows where well was empty
    IncucyteMetaData.dropna(subset=['cell_line_modifications'], inplace=True)

    # remove nonsense rows where cell_line = None:

    IncucyteMetaDataUpdated = pd.DataFrame()

    IncucyteMetaDataUpdated = incu_mysql_upload_functions.removeNonsenseCellLines(IncucyteMetaData,IncucyteMetaDataUpdated)

    # create empty series
    SeriesOfConditions = pd.Series([],dtype=pd.StringDtype())


    SeriesOfConditions = incu_mysql_upload_functions.assignConditionID(IncucyteMetaDataUpdated, SeriesOfConditions)


    # add condition IDs to df
    IncucyteMetaDataUpdated['condition_ID'] = SeriesOfConditions


    # move the condition ID column to head of list using index, pop and insert
    # get a list of columns
    ListOfColumns = list(IncucyteMetaDataUpdated)
    ListOfColumns.insert(0, ListOfColumns.pop(ListOfColumns.index('condition_ID')))

    # save new column ordering
    NewConditionsDF = IncucyteMetaDataUpdated.loc[:, ListOfColumns] 

    # remove nans and replace with None
    NewConditionsDF = NewConditionsDF.replace(np.nan,None)

    # remove condition duplicates for all columns
    NewConditionsDF.drop_duplicates(ignore_index=True,inplace = True)

    # get length of dataframe
    TableSize = NewConditionsDF.shape[0]
    
    # insert query
    sqlInsert = """INSERT INTO BAF_project.Incucyte_Condition_Table (condition_ID, cell_line, 
                        cell_line_modifications, cell_number, treatment_0, concentration_0, treatment_timepoint_0, 
                        treatment_duration_0,treatment_1, concentration_1, treatment_timepoint_1, 
                        treatment_duration_1,treatment_2, concentration_2, treatment_timepoint_2, 
                        treatment_duration_2,treatment_3, concentration_3, treatment_timepoint_3, 
                        treatment_duration_3) 
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)"""

    # insert values from condition_df_final row-wise into the MySQL table
    for i in range(TableSize):
        cursor.execute(sqlInsert,tuple(NewConditionsDF.iloc[i]))


# commit to the db
connect.commit()


###########################################################################################################################################################################

# Incucyte Metadata Table

# get cell line and treatment info and condition IDs from existing Incucyte Condition ID table from MySQL
query2 = """SELECT * FROM BAF_project.Incucyte_Condition_Table"""
cursor.execute(query2)
IncucyteConditionTableTuple = cursor.fetchall()

# create dataframe of IncucyteConditionTableTuple
ConditionTableContents = pd.DataFrame(IncucyteConditionTableTuple, columns=['condition_ID', 'cell_line', 
                    'cell_line_modifications', 'cell_number', 'treatment_0', 'concentration_0', 'treatment_timepoint_0', 
                    'treatment_duration_0','treatment_1', 'concentration_1', 'treatment_timepoint_1', 
                    'treatment_duration_1','treatment_2', 'concentration_2', 'treatment_timepoint_2', 
                    'treatment_duration_2','treatment_3', 'concentration_3', 'treatment_timepoint_3', 
                    'treatment_duration_3'])

# insert query
sqlInsert3 = """INSERT INTO BAF_project.Incucyte_Meta_Data (project_ID,plate_ID,well_ID,condition_ID) 
                VALUES (%s,%s,%s,%s)"""

#create empty dataframe 
IncucyteMetaData = pd.DataFrame()

IncucyteMetaData = incu_mysql_upload_functions.loadIncucyteMetaData2(ListofIncucytePlatelayouts, IncucyteMetaData)

#remove nans and replace with None
IncucyteMetaData = IncucyteMetaData.replace(np.nan,None)

#reset index numbering
IncucyteMetaData=IncucyteMetaData.reset_index()

#remove unnecessary column
IncucyteMetaData= IncucyteMetaData.drop(['index'], axis =1)

MetaDataTable = incu_mysql_upload_functions.createMetaDataTable(IncucyteMetaData,ConditionTableContents)


MetaDataTableExistsInMySQL =incu_mysql_upload_functions.checkTableExists('Incucyte_Meta_Data',cursor)


#check if Meta Data Table already exists in MySQL

if MetaDataTableExistsInMySQL == False:

    create_statement2=  """CREATE TABLE BAF_project.Incucyte_Meta_Data (
        project_ID varchar(45) NOT NULL,
        plate_ID varchar(25) NOT NULL,
        well_ID varchar(45) NOT NULL,
        condition_ID varchar(45) NOT NULL,
        PRIMARY KEY (project_ID,plate_ID,well_ID))"""
    cursor.execute(create_statement2)

    # insert values row-wise into MySQL table
    TableSize = MetaDataTable.shape[0]
    for i in range(TableSize):
        cursor.execute(sqlInsert3,tuple(MetaDataTable.loc[i]))


if MetaDataTableExistsInMySQL == True:

    # load existing metadata from database
    query3 = """SELECT * FROM BAF_project.Incucyte_Meta_Data"""

    cursor.execute(query3)

    MetaTuple = cursor.fetchall()

    ExistingMetaDataTableContents = pd.DataFrame(MetaTuple, columns=['project_ID','plate_ID','well_ID','condition_ID'])

    # merge existing metadata with new metadata
    UpdatedMetaDataTable = pd.merge(MetaDataTable,ExistingMetaDataTableContents, on=['project_ID','plate_ID','well_ID','condition_ID'], how="left", indicator = True)

    UpdatedMetaDataTable = UpdatedMetaDataTable.loc[UpdatedMetaDataTable['_merge'] == 'left_only'].drop(columns=['_merge'])

    # update metadata table in database; insert values row-wise into MySQL table
    TableSize = UpdatedMetaDataTable.shape[0]
    for i in range(TableSize):
        cursor.execute(sqlInsert3,tuple(UpdatedMetaDataTable.iloc[i]))

# commit to the db
connect.commit()

cursor.close()
